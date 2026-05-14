# utils/excel_writer.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SHEET_TITLES = [
    "Sheet1-Model1",
    "Sheet2-Model2",
    "Sheet3-Model3",
    "Sheet4-Consolidated",
]


def _to_df(results: list) -> pd.DataFrame:
    rows = []
    for item in results:
        (rows.extend(item) if isinstance(item, list) else rows.append(item))
    return pd.DataFrame(rows) if rows else pd.DataFrame({"note": ["No data"]})


def _fill_sheet(ws, df: pd.DataFrame, title: str) -> None:
    ws.title = title
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for i in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 28


def export_excel(sheet1, sheet2, sheet3, sheet4, out_path: str) -> str:
    wb = Workbook()
    datasets = [sheet1, sheet2, sheet3, sheet4]
    for idx, (data, title) in enumerate(zip(datasets, _SHEET_TITLES)):
        ws = wb.active if idx == 0 else wb.create_sheet()
        _fill_sheet(ws, _to_df(data), title)
    wb.save(out_path)
    return out_path
